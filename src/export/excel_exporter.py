"""Excel台帳出力モジュール

DBのデータを処理台帳・仕入台帳のExcelファイルに書き出す。
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.common.config import get_setting, BASE_DIR
from src.common.database import get_connection
from src.common.logger import get_logger

logger = get_logger(__name__)

# スタイル定義
HEADER_FONT = Font(name="Yu Gothic", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CELL_FONT = Font(name="Yu Gothic", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 在庫ステータス → 日本語ラベル
STOCK_STATUS_LABELS = {
    "AVAILABLE": "在庫あり",
    "UNAVAILABLE": "在庫なし",
    "BACKORDER": "取り寄せ可",
    "UNKNOWN": "不明",
    "ERROR": "エラー",
}

# 在庫ステータス別の塗りつぶし色
STATUS_FILLS = {
    "在庫あり": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "在庫なし": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "取り寄せ可": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "不明": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "エラー": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

# 保留バケットステータス → 日本語ラベル
BUCKET_STATUS_LABELS = {
    "ACTIVE": "積上中",
    "THRESHOLD_REACHED": "閾値到達",
    "CLEARED": "クリア",
}


def _stock_label(status: str) -> str:
    """在庫ステータスを日本語ラベルに変換する。"""
    return STOCK_STATUS_LABELS.get(status, status)


def _bucket_label(status: str) -> str:
    """バケットステータスを日本語ラベルに変換する。"""
    return BUCKET_STATUS_LABELS.get(status, status)


def _apply_header_style(ws, row: int, col_count: int) -> None:
    """ヘッダー行にスタイルを適用する。"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_cell_style(ws, row: int, col_count: int) -> None:
    """データ行にスタイルを適用する。"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = CELL_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")


def export_processing_ledger(output_path: str | None = None) -> str:
    """処理台帳（①）をExcelに出力する。

    1商品1行。各仕入先の在庫結果を横に展開する。

    Returns:
        出力ファイルパス
    """
    if output_path is None:
        output_dir = os.path.join(
            BASE_DIR, get_setting("storage", "output_dir", default="data/output")
        )
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(output_dir, f"processing_ledger_{timestamp}.xlsx")

    # 仕入先一覧を取得（列ヘッダー用）
    with get_connection() as conn:
        suppliers = conn.execute(
            "SELECT id, supplier_code, supplier_name FROM suppliers ORDER BY priority"
        ).fetchall()
        supplier_list = [dict(s) for s in suppliers]

    # 固定列ヘッダー
    fixed_headers = [
        "No.", "管理番号", "受信日", "商品名", "商品コード",
        "金額", "数量", "自家在庫", "状態", "予定仕入先", "振分日",
    ]

    # 仕入先別の在庫列
    supplier_headers = [s["supplier_name"] for s in supplier_list]
    all_headers = fixed_headers + supplier_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "処理台帳"

    # ヘッダー書き込み
    for col, header in enumerate(all_headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(all_headers))

    # データ取得
    with get_connection() as conn:
        items = conn.execute(
            """
            SELECT oi.*,
                   m.received_at AS mail_received_at,
                   m.subject AS mail_subject,
                   ps.supplier_name AS planned_supplier_name
            FROM order_items oi
            LEFT JOIN messages m ON oi.message_id = m.id
            LEFT JOIN suppliers ps ON oi.planned_supplier_id = ps.id
            ORDER BY oi.id DESC
            """
        ).fetchall()

    for row_idx, item in enumerate(items, 2):
        item = dict(item)
        is_self_stock = item["current_status"] == "SELF_STOCK"

        row_data = [
            row_idx - 1,
            item["order_number"],
            item["mail_received_at"] or "",
            item["product_name"],
            item["product_code_normalized"],
            item["amount"],
            item["quantity"],
            "○" if is_self_stock else "",
            item["current_status"],
            item["planned_supplier_name"] or "",
            item["assigned_at"] or "",
        ]

        # 各仕入先の最新在庫結果を取得（日本語ラベルで表示）
        with get_connection() as conn:
            for sup in supplier_list:
                result = conn.execute(
                    """
                    SELECT availability_status FROM scrape_results
                    WHERE order_item_id = ? AND supplier_id = ?
                    ORDER BY id DESC LIMIT 1
                    """,
                    (item["id"], sup["id"]),
                ).fetchone()

                status_text = ""
                if result:
                    status_text = _stock_label(result["availability_status"])
                row_data.append(status_text)

        # 書き込み
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
        _apply_cell_style(ws, row_idx, len(all_headers))

        # 在庫ステータスに色をつける（日本語ラベルでマッチ）
        for col_offset, sup in enumerate(supplier_list):
            col_num = len(fixed_headers) + col_offset + 1
            cell = ws.cell(row=row_idx, column=col_num)
            fill = STATUS_FILLS.get(str(cell.value))
            if fill:
                cell.fill = fill

    # 列幅調整
    column_widths = {
        1: 6, 2: 16, 3: 18, 4: 30, 5: 18,
        6: 10, 7: 6, 8: 10, 9: 12, 10: 18, 11: 18,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for i in range(len(supplier_list)):
        ws.column_dimensions[get_column_letter(len(fixed_headers) + i + 1)].width = 16

    # ウィンドウ枠固定
    ws.freeze_panes = "A2"

    # オートフィルタ
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{len(items) + 1}"

    wb.save(output_path)
    logger.info("処理台帳出力完了: %s (%d件)", output_path, len(items))
    return output_path


def export_supplier_ledger(output_path: str | None = None) -> str:
    """仕入台帳（②）をExcelに出力する。

    1仕入先1行。保留合計金額・件数・経過日数・全保留商品を表示。

    Returns:
        出力ファイルパス
    """
    if output_path is None:
        output_dir = os.path.join(
            BASE_DIR, get_setting("storage", "output_dir", default="data/output")
        )
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(output_dir, f"supplier_ledger_{timestamp}.xlsx")

    # まず全バケットの保留商品数の最大値を取得（動的列数）
    with get_connection() as conn:
        max_items_row = conn.execute(
            """
            SELECT MAX(cnt) AS max_cnt FROM (
                SELECT COUNT(*) AS cnt FROM hold_items GROUP BY hold_bucket_id
            )
            """
        ).fetchone()
    max_hold_items = (max_items_row["max_cnt"] or 0) if max_items_row else 0
    max_hold_items = max(max_hold_items, 1)  # 最低1列

    # 固定ヘッダー + 動的な保留商品列
    fixed_headers = [
        "No.", "仕入先コード", "仕入先名", "カテゴリ",
        "保留合計金額", "保留件数", "最経過日数",
        "保留上限金額", "ステータス",
    ]
    item_headers = [f"保留商品{i+1}" for i in range(max_hold_items)]
    all_headers = fixed_headers + item_headers

    wb = Workbook()
    ws = wb.active
    ws.title = "仕入台帳"

    for col, header in enumerate(all_headers, 1):
        ws.cell(row=1, column=col, value=header)
    _apply_header_style(ws, 1, len(all_headers))

    with get_connection() as conn:
        buckets = conn.execute(
            """
            SELECT hb.*, s.supplier_code, s.supplier_name, s.category, s.hold_limit_amount
            FROM hold_buckets hb
            JOIN suppliers s ON hb.supplier_id = s.id
            ORDER BY s.priority
            """
        ).fetchall()

    for row_idx, bucket in enumerate(buckets, 2):
        bucket = dict(bucket)

        # 経過日数計算
        elapsed_days = 0
        if bucket["oldest_item_date"]:
            try:
                oldest = datetime.strptime(bucket["oldest_item_date"], "%Y-%m-%d %H:%M:%S")
                elapsed_days = (datetime.now() - oldest).days
            except ValueError:
                pass

        row_data = [
            row_idx - 1,
            bucket["supplier_code"],
            bucket["supplier_name"],
            "EC" if bucket["category"] == "ec" else "店頭",
            bucket["total_amount"],
            bucket["item_count"],
            elapsed_days,
            bucket["hold_limit_amount"],
            _bucket_label(bucket["status"]),
        ]

        # 保留中の全商品（件数制限なし）
        with get_connection() as conn:
            hold_items = conn.execute(
                """
                SELECT oi.product_name, oi.product_code_normalized, oi.amount,
                       m.received_at
                FROM hold_items hi
                JOIN order_items oi ON hi.order_item_id = oi.id
                LEFT JOIN messages m ON oi.message_id = m.id
                WHERE hi.hold_bucket_id = ?
                ORDER BY hi.assigned_at ASC
                """,
                (bucket["id"],),
            ).fetchall()

        for i in range(max_hold_items):
            if i < len(hold_items):
                hi = hold_items[i]
                name = hi["product_name"] or ""
                code = hi["product_code_normalized"] or ""
                amt = hi["amount"] or 0
                row_data.append(f"{name} ({code}) ¥{amt:,}")
            else:
                row_data.append("")

        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
        _apply_cell_style(ws, row_idx, len(all_headers))

    # 列幅調整
    widths = {1: 6, 2: 16, 3: 20, 4: 8, 5: 14, 6: 10, 7: 12, 8: 14, 9: 12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for i in range(max_hold_items):
        ws.column_dimensions[get_column_letter(len(fixed_headers) + i + 1)].width = 35

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{len(buckets) + 1}"

    wb.save(output_path)
    logger.info("仕入台帳出力完了: %s (%d件)", output_path, len(buckets))
    return output_path


def export_all() -> tuple[str, str]:
    """処理台帳と仕入台帳を両方出力する。"""
    p = export_processing_ledger()
    s = export_supplier_ledger()
    return p, s
