"""URL指定型ファイル読込モジュール（機能②）

仕様書 TAB2 NO3 機能②:
  事前に列「ISBN・ASINコード」「URL」を含むファイルを読み込み、
  NO1で処理したコードと照合してスクレイピング先URLを決定する。

対応フォーマット:
  - .xlsx / .xlsm  : openpyxl で読込
  - .tsv          : タブ区切り
  - .csv          : カンマ区切り

ファイル形式（13列、1行目はヘッダー）:
  1: ISBN（照合キー）
  2: 書名
  3: 出版社
  4: 著者
  5: 発売年
  6: 定価
  7: ページURL（検索結果一覧）
  8: フィールド1_テキスト（商品名）
  9: フィールド1_リンク（商品詳細URL）← スクレイピング先として使用
  10: 価格 ← 最安値比較に使用
  11: フィールド2
  12: フィールド3
  13: 現在の時刻

同一ISBNに複数行がある場合、価格が最安の行を優先する（is_cheapest=1）。
"""

import csv
import os
import re

from src.common.config import get_setting, BASE_DIR
from src.common.database import get_connection
from src.common.logger import get_logger
from src.parser.code_normalizer import normalize_code

logger = get_logger(__name__)


def _parse_price(price_str: str) -> int:
    """価格文字列を整数に変換する。カンマ・通貨記号を除去。"""
    if not price_str or not price_str.strip():
        return 0
    cleaned = re.sub(r"[^\d]", "", price_str.strip())
    if not cleaned:
        return 0
    try:
        return int(cleaned)
    except ValueError:
        return 0


def _cell_str(value) -> str:
    """セル値を strip 済みの文字列に正規化する。"""
    if value is None:
        return ""
    return str(value).strip()


def _build_row(values: list) -> tuple | None:
    """13列の生データから、DB INSERT 用のタプルを構築する。

    item_url（9列目）が空、または商品コードが正規化できない行は None を返す。
    """
    if len(values) < 9:
        return None

    raw_code = _cell_str(values[0])
    if not raw_code:
        return None

    code = normalize_code(raw_code)
    if not code:
        return None

    item_url = _cell_str(values[8])
    if not item_url:
        return None

    book_title = _cell_str(values[1]) if len(values) > 1 else ""
    publisher = _cell_str(values[2]) if len(values) > 2 else ""
    author = _cell_str(values[3]) if len(values) > 3 else ""
    pub_year = _cell_str(values[4]) if len(values) > 4 else ""
    list_price = _parse_price(_cell_str(values[5])) if len(values) > 5 else 0
    page_url = _cell_str(values[6]) if len(values) > 6 else ""
    item_text = _cell_str(values[7]) if len(values) > 7 else ""
    price = _parse_price(_cell_str(values[9])) if len(values) > 9 else 0
    field2 = _cell_str(values[10]) if len(values) > 10 else ""
    field3 = _cell_str(values[11]) if len(values) > 11 else ""
    scraped_at = _cell_str(values[12]) if len(values) > 12 else ""

    return (
        code, book_title, publisher, author, pub_year,
        list_price, page_url, item_text, item_url, price,
        field2, field3, scraped_at,
    )


def _read_url_csv_rows(file_path: str) -> list[tuple]:
    """URL指定型ファイルから INSERT 用タプルのリストを返す。"""
    ext = os.path.splitext(file_path)[1].lower()
    rows: list[tuple] = []

    if ext in (".xlsx", ".xlsm"):
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        try:
            ws = wb.active
            iter_rows = ws.iter_rows(values_only=True)
            next(iter_rows, None)  # ヘッダー行を読み飛ばし
            for raw in iter_rows:
                if not raw:
                    continue
                built = _build_row(list(raw))
                if built is not None:
                    rows.append(built)
        finally:
            wb.close()
        return rows

    delim = "\t" if ext == ".tsv" else ","
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f, delimiter=delim)
        next(reader, None)
        for raw in reader:
            built = _build_row(raw)
            if built is not None:
                rows.append(built)
    return rows


def import_url_csv(csv_path: str | None = None) -> int:
    """URL指定型ファイルを読み込み、url_csv_entriesテーブルに洗い替えする。

    同一商品コードで複数行がある場合、最安値の行に is_cheapest=1 を付与する。

    Args:
        csv_path: ファイルパス。省略時は設定 url_csv.csv_path から取得。

    Returns:
        取込件数
    """
    if csv_path is None:
        rel_path = get_setting("url_csv", "csv_path", default="")
        if not rel_path:
            logger.info("URL指定型ファイルのパスが設定されていません（url_csv.csv_path）")
            return 0
        csv_path = os.path.join(BASE_DIR, rel_path)

    if not os.path.exists(csv_path):
        logger.warning("URL指定型ファイルが見つかりません: %s", csv_path)
        return 0

    try:
        rows = _read_url_csv_rows(csv_path)
    except Exception as e:
        logger.error("URL指定型ファイル読込エラー (%s): %s", csv_path, e)
        return 0

    if not rows:
        logger.warning("URL指定型CSVに有効なデータがありません: %s", csv_path)
        return 0

    # 同一コードごとに最安値を特定
    code_best: dict[str, int] = {}  # {product_code: best_price}
    for r in rows:
        code = r[0]
        price = r[9]  # price列
        if price <= 0:
            continue
        if code not in code_best or price < code_best[code]:
            code_best[code] = price

    # 洗い替え（DELETE → INSERT）
    with get_connection() as conn:
        conn.execute("DELETE FROM url_csv_entries")
        for r in rows:
            code = r[0]
            price = r[9]
            # 最安値判定: 価格が一致する場合のみ is_cheapest=1（同額なら先行行が優先）
            is_cheapest = 0
            if price > 0 and code in code_best and price == code_best[code]:
                is_cheapest = 1
                # 一度フラグを立てたら同一コードの次の同額行はフラグを立てない
                del code_best[code]

            conn.execute(
                """
                INSERT INTO url_csv_entries (
                    product_code, book_title, publisher, author, pub_year,
                    list_price, page_url, item_text, item_url, price,
                    field2, field3, scraped_at, is_cheapest
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                r + (is_cheapest,),
            )

    cheapest_count = sum(1 for r in rows if r[0] in {} or True)  # 再計算
    with get_connection() as conn:
        cheapest_count = conn.execute(
            "SELECT COUNT(*) as cnt FROM url_csv_entries WHERE is_cheapest = 1"
        ).fetchone()["cnt"]
        total_count = conn.execute(
            "SELECT COUNT(*) as cnt FROM url_csv_entries"
        ).fetchone()["cnt"]

    logger.info(
        "URL指定型CSV取込完了: 全%d件, ユニーク最安値%d件",
        total_count, cheapest_count,
    )
    return total_count


def lookup_cheapest_url(product_code: str) -> dict | None:
    """商品コードから最安値のURL情報を取得する。

    Args:
        product_code: 正規化済み商品コード（ISBN-13 or ASIN）

    Returns:
        {"item_url": str, "price": int, "item_text": str, ...} or None
    """
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT item_url, price, item_text, book_title, publisher, author
            FROM url_csv_entries
            WHERE product_code = ? AND is_cheapest = 1
            LIMIT 1
            """,
            (product_code,),
        ).fetchone()

    if row is None:
        return None

    return {
        "item_url": row["item_url"],
        "price": row["price"],
        "item_text": row["item_text"],
        "book_title": row["book_title"],
        "publisher": row["publisher"],
        "author": row["author"],
    }


def has_url_csv_entry(product_code: str) -> bool:
    """指定商品コードがURL指定型CSVに存在するかを確認する。"""
    with get_connection() as conn:
        row = conn.execute(
            "SELECT 1 FROM url_csv_entries WHERE product_code = ? LIMIT 1",
            (product_code,),
        ).fetchone()
    return row is not None
