"""自家在庫照合モジュール

自社の在庫ファイル（xlsx または csv）を読み込み、
order_itemsの商品コードと突き合わせる。
自家在庫がある商品はスクレイピング対象から外す。
"""

import csv
import os

from src.common.config import get_setting, BASE_DIR
from src.common.database import get_connection
from src.common.enums import OrderStatus
from src.common.logger import get_logger
from src.parser.code_normalizer import normalize_code

logger = get_logger(__name__)

# ヘッダー名 → 役割 のゆらぎ吸収
_CODE_HEADERS = {"isbn", "jan", "コード", "商品コード", "isbnコード"}
_QTY_HEADERS = {"冊数", "在庫数", "数量", "qty"}


def _read_self_stock_rows(file_path: str) -> list[tuple[str, int]]:
    """自家在庫ファイルから (商品コード, 在庫数) のリストを返す。

    .xlsx / .xlsm: openpyxl で読み、ヘッダー名から ISBN/冊数 列を特定。
    .csv (テスト等の互換用): 1列目=商品コード, 2列目=在庫数 の旧形式。
    """
    ext = os.path.splitext(file_path)[1].lower()
    rows: list[tuple[str, int]] = []

    if ext in (".xlsx", ".xlsm"):
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        try:
            ws = wb.active
            iter_rows = ws.iter_rows(values_only=True)
            header = next(iter_rows, None)
            if header is None:
                return rows

            code_idx: int | None = None
            qty_idx: int | None = None
            for i, h in enumerate(header):
                if h is None:
                    continue
                hs = str(h).strip().lower()
                if code_idx is None and hs in _CODE_HEADERS:
                    code_idx = i
                elif qty_idx is None and hs in _QTY_HEADERS:
                    qty_idx = i

            if code_idx is None:
                logger.warning(
                    "自家在庫xlsxにISBN/商品コード列が見つかりません: %s", file_path
                )
                return rows

            for row in iter_rows:
                if not row or code_idx >= len(row):
                    continue
                raw = row[code_idx]
                if raw is None:
                    continue
                raw_str = str(raw).strip()
                if not raw_str:
                    continue
                code = normalize_code(raw_str)
                if not code:
                    continue
                qty = 0
                if qty_idx is not None and qty_idx < len(row):
                    v = row[qty_idx]
                    if v is not None:
                        try:
                            qty = int(float(v))
                        except (ValueError, TypeError):
                            qty = 0
                rows.append((code, qty))
        finally:
            wb.close()
        return rows

    # 旧CSV形式（互換）
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header is None:
            return rows
        for row in reader:
            if len(row) < 1 or not row[0].strip():
                continue
            code = normalize_code(row[0].strip())
            qty = 0
            if len(row) >= 2:
                try:
                    qty = int(row[1].strip())
                except ValueError:
                    qty = 0
            if code:
                rows.append((code, qty))
    return rows


def import_self_stock_csv(csv_path: str | None = None) -> int:
    """自家在庫ファイルを読み込み、self_stockテーブルに洗い替えする。

    対応フォーマット:
      - .xlsx / .xlsm: ヘッダー名 "ISBN"・"冊数"（または同義語）から列を特定
      - .csv: 旧形式（1列目=商品コード, 2列目=在庫数）

    Args:
        csv_path: ファイルパス。省略時は設定 self_stock.csv_path から取得。

    Returns:
        取込件数
    """
    if csv_path is None:
        rel_path = get_setting("self_stock", "csv_path", default="data/csv/self_stock.xlsx")
        csv_path = os.path.join(BASE_DIR, rel_path)

    if not os.path.exists(csv_path):
        logger.warning("自家在庫ファイルが見つかりません: %s", csv_path)
        return 0

    try:
        rows = _read_self_stock_rows(csv_path)
    except Exception as e:
        logger.error("自家在庫ファイル読込エラー (%s): %s", csv_path, e)
        return 0

    # 洗い替え（DELETE → INSERT）
    with get_connection() as conn:
        conn.execute("DELETE FROM self_stock")
        conn.executemany(
            "INSERT INTO self_stock (product_code, stock_qty) VALUES (?, ?)",
            rows,
        )

    logger.info("自家在庫CSV取込完了: %d件", len(rows))
    return len(rows)


def check_self_stock() -> int:
    """order_items(PENDING)の商品コードを自家在庫と照合し、
    一致した商品のステータスをSELF_STOCKに更新する。

    Returns:
        自家在庫ヒット件数
    """
    with get_connection() as conn:
        # 自家在庫に存在する商品コードの集合を取得
        stock_rows = conn.execute(
            "SELECT product_code FROM self_stock WHERE stock_qty > 0"
        ).fetchall()
        stock_codes = {row["product_code"] for row in stock_rows}

        if not stock_codes:
            logger.info("自家在庫データがないため照合をスキップします")
            return 0

        # PENDING状態の商品を取得
        pending_items = conn.execute(
            "SELECT id, product_code_normalized FROM order_items WHERE current_status = ?",
            (OrderStatus.PENDING.value,),
        ).fetchall()

        hit_count = 0
        for item in pending_items:
            if item["product_code_normalized"] in stock_codes:
                conn.execute(
                    "UPDATE order_items SET current_status = ? WHERE id = ?",
                    (OrderStatus.SELF_STOCK.value, item["id"]),
                )
                hit_count += 1

    if hit_count > 0:
        logger.info("自家在庫照合: %d件が自家在庫にヒット", hit_count)
    else:
        logger.info("自家在庫照合: ヒットなし")

    return hit_count
